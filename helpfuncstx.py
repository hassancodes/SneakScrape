import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from fake_useragent import UserAgent
from random import randint
import random

####################################################################
def chooseOption(ls):
    print("Choose 1 or 2 \n1.) Start scraping from index 1 \n2.) Start scrapping from a specific index.\nEnter Below:")
    try:
        inp = int(input())
        if inp==1:
            print("starting script from index 1")
            # we will subtract index-1 later in stockx.py so we can get the correct index
            return 1
        elif inp == 2:
            try:
                print("input the index number from excel: ")
                rindex = input()
                index = int(rindex)
                if index<len(ls) and index>0:
                    print(f"Starting scraping from index {index}")
                    # this index can be used in stockx.py
                    return index
                else:
                    print("invalid input: Enter a valid Index!")
            except:
                print("invalid input: Enter a valid Index!")
        else:
            print("invalid input")
            return None
    except:
        print("invalid input! You can only choose 1 or 2")



####################################################################
# function to generate the main url.
# this functions converts - ""adidas Crazy BYW X 2.0 Ubiq" to "adidas-crazy-byw-x-2-ubiq"
# later we will connect this with stockx.com
def createUrl(SneakerName):
    ls = SneakerName.lower().split(" ")
    parsedls =[]
    for i in ls:
            if i.replace('.','').isdigit():
                parsedls.append(str(int(i.split(".")[0])))

            else:
                parsedls.append(i)
    endpoint = "-".join(parsedls)
    #
    return endpoint

######################################################################
######################################################################

# function for getting style codes from the excel sheet
def getStyleCode():
    styleCodes = load_workbook("StyleCodes.xlsx")
    sheet = styleCodes.active
    rawsdList = [x.value for x in sheet["A"]]
# main style code list
    StyleCodeList = rawsdList[1:]
    # print("Total Style codes: ",len(StyleCodeList))
    return StyleCodeList

#######################################################################
#######################################################################

# function for generating fake useragent
def ua():

    ua = UserAgent()
    return ua.chrome

def spacefunc():
    print("\n")


# getting the divs
def parseDiv(ls):
    for i in range(len(ls)):
        stripped = str(ls[i]).strip()[8:].strip()
        if stripped.startswith("window.__APOLLO_STATE_"):
            return ls[i]
            break
        else:
            pass

#######################################################################
#######################################################################

def parseIp():
    proxyList = []
    with open("ip.txt") as ip:
        for i in ip:
            iplist = i.split(":")
            readyProxy = "".join((iplist[0],":",iplist[1]))
            proxyList.append(readyProxy)


    random.shuffle(proxyList)
    return proxyList




# import requests
# url = 'https://api.myip.com'
# wIpList = parseIp()
# counter =
# proxy = {
# "http":wIpList[counter],
#  "https":wIpList[counter]
# }
#
# response = requests.get(url, proxies = proxy)
# print(response.json())



def randomIp(counter):
    ipList = parseIp()
    wIpList = ipList.copy()

    if counter<len(wIpList):
        # print(counter , ":<", wIpList[counter])
        return wIpList[counter]

    elif counter > 2*(len(wIpList)) :
        r = random.randint(0,len(wIpList)-1)
        # print(r , wIpList[r])
        # print("random: ", wIpList[r])
        return wIpList[r]

    elif counter==len(wIpList) or counter < 2*(len(wIpList)):
        # print(counter , ":=>", wIpList[len(wIpList)-counter] )
        return wIpList[len(wIpList)-counter]


############################Function for checking successrate of a script####################################

def successRate(ls):
    totalEntries = len(ls)
    styleCodes = load_workbook("StyleCodes.xlsx")
    sheet = styleCodes.active
    rawsdList = [x.value for x in sheet["B"]]
# main style code list
    unfilterls = rawsdList[1:]
    filteredls = list(filter(lambda x:x!=None and x!="Not found",unfilterls))
    print("Total Shoes Entries: ", totalEntries)
    print("Successfully added Sneaker Name: ", len(filteredls))
    return f"Success rate is: {round((len(filteredls)/totalEntries)*100 , 2)}%"

##################################
